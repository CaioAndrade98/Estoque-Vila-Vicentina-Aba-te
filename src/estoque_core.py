from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Dict, List
from .config import ARQUIVO_DADOS, BACKUP_DIR, ARQUIVO_HISTORICO

import csv
from pathlib import Path



class ProdutoNaoEncontrado(Exception):
    """Produto não encontrado no estoque."""


class EstoqueInsuficiente(Exception):
    """Quantidade insuficiente em estoque."""


class ProdutoDuplicado(Exception):
    """Já existe um produto com esse nome (normalizado)."""


def _normalizar_nome(nome: str) -> str:
    return " ".join(nome.strip().lower().split())


def _carregar_produtos() -> List[Dict[str, Any]]:
    if not ARQUIVO_DADOS.exists():
        return []
    try:
        with ARQUIVO_DADOS.open("r", encoding="utf-8") as f:
            dados = json.load(f)
        return dados if isinstance(dados, list) else []
    except Exception:
        return []


def _salvar_produtos(produtos: List[Dict[str, Any]]) -> None:
    ARQUIVO_DADOS.parent.mkdir(parents=True, exist_ok=True)

    with ARQUIVO_DADOS.open("w", encoding="utf-8") as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)

    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        carimbo = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_file = BACKUP_DIR / f"dados_backup_{carimbo}.json"
        with backup_file.open("w", encoding="utf-8") as bf:
            json.dump(produtos, bf, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _registrar_movimento(
    produto_id: int,
    nome: str,
    delta: float,
    estoque_antes: float,
    estoque_depois: float,
) -> None:
    """
    Registra um movimento em JSON Lines (%APPDATA%\\EstoqueONG\\historico\\movimentos.jsonl).
    Um JSON por linha para ser fácil de ler/exportar depois.
    """
    try:
        evento = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "produto_id": int(produto_id),
            "nome": str(nome),
            "delta": float(delta),  # +entrada / -saida
            "estoque_antes": float(estoque_antes),
            "estoque_depois": float(estoque_depois),
        }

        # garante pasta (por segurança)
        ARQUIVO_HISTORICO.parent.mkdir(parents=True, exist_ok=True)

        with ARQUIVO_HISTORICO.open("a", encoding="utf-8") as f:
            f.write(json.dumps(evento, ensure_ascii=False) + "\n")
    except Exception:
        # Histórico nunca pode quebrar o app
        pass


def listar_produtos() -> List[Dict[str, Any]]:
    """Retorna todos os produtos (lista de dicts) do JSON."""
    return _carregar_produtos()


def _gerar_proximo_id(produtos: List[Dict[str, Any]]) -> int:
    if not produtos:
        return 1
    return max(int(p.get("id", 0)) for p in produtos) + 1


def criar_produto(nome: str, unidade: str, estoque_minimo: float) -> Dict[str, Any]:
    """Cria produto (estoque_atual=0.0) e persiste."""
    nome = str(nome).strip()
    unidade = str(unidade).strip()
    if not nome:
        raise ValueError("Nome não pode ficar vazio.")
    if not unidade:
        raise ValueError("Unidade não pode ficar vazia.")

    try:
        estoque_min = float(str(estoque_minimo).replace(",", "."))
    except Exception:
        raise ValueError("Estoque mínimo inválido.")
    if estoque_min < 0:
        raise ValueError("Estoque mínimo inválido.")

    produtos = _carregar_produtos()
    nome_norm = _normalizar_nome(nome)
    for p in produtos:
        if _normalizar_nome(str(p.get("nome", ""))) == nome_norm:
            raise ProdutoDuplicado("Produto já existe com esse nome.")

    novo = {
        "id": _gerar_proximo_id(produtos),
        "nome": nome,
        "unidade": unidade,
        "estoque_atual": 0.0,
        "estoque_minimo": float(estoque_min),
    }
    produtos.append(novo)
    _salvar_produtos(produtos)
    return novo


def move_stock_by_id(produto_id: int, delta: float) -> Dict[str, Any]:
    """Movimenta estoque_atual por ID. delta positivo=entrada; negativo=saída."""
    try:
        pid = int(produto_id)
    except Exception:
        raise ProdutoNaoEncontrado("ID inválido.")

    try:
        d = float(delta)
    except Exception:
        raise ValueError("Quantidade inválida.")

    produtos = _carregar_produtos()
    for p in produtos:
        if int(p.get("id", 0)) == pid:
            atual = float(p.get("estoque_atual", 0.0))
            novo = atual + d
            if novo < 0:
                raise EstoqueInsuficiente(f"Estoque insuficiente para '{p.get('nome', '')}'.")
            p["estoque_atual"] = float(novo)
            _salvar_produtos(produtos)

            _registrar_movimento(
                produto_id=pid,
                nome=str(p.get("nome", "")),
                delta=float(d),
                estoque_antes=float(atual),
                estoque_depois=float(novo),
)

            return p


    raise ProdutoNaoEncontrado(f"Produto com id {pid} não encontrado.")


def produtos_abaixo_minimo() -> List[Dict[str, Any]]:
    """Produtos cujo estoque_atual está abaixo do estoque_minimo."""
    produtos = _carregar_produtos()
    return [
        p for p in produtos
        if float(p.get("estoque_atual", 0.0)) < float(p.get("estoque_minimo", 0.0))
    ]

def listar_movimentos(limite: int | None = None) -> list[dict]:
    """
    Retorna os movimentos do histórico (mais recentes por último).
    Se limite for informado, retorna apenas os últimos N movimentos.
    """
    if not ARQUIVO_HISTORICO.exists():
        return []

    movimentos: list[dict] = []

    try:
        with ARQUIVO_HISTORICO.open("r", encoding="utf-8") as f:
            for linha in f:
                linha = linha.strip()
                if not linha:
                    continue
                try:
                    movimentos.append(json.loads(linha))
                except Exception:
                    continue
    except Exception:
        return []

    if limite is not None and limite > 0:
        return movimentos[-limite:]

    return movimentos

def exportar_movimentos_csv(caminho_csv: str | Path, limite: int | None = None) -> Path:
    """
    Exporta o histórico de movimentos para CSV.
    Retorna o Path do arquivo gerado.
    """
    caminho = Path(caminho_csv)

    movimentos = listar_movimentos(limite=limite)

    # garante pasta do arquivo
    caminho.parent.mkdir(parents=True, exist_ok=True)

    with caminho.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["ts", "produto_id", "nome", "delta", "estoque_antes", "estoque_depois"])

        for m in movimentos:
            w.writerow([
                m.get("ts", ""),
                m.get("produto_id", ""),
                m.get("nome", ""),
                m.get("delta", ""),
                m.get("estoque_antes", ""),
                m.get("estoque_depois", ""),
            ])

    return caminho

def exportar_movimentos_xlsx(caminho_xlsx, movimentos):
    """
    Exporta uma lista de movimentos para um arquivo Excel (.xlsx)
    com formatação básica (cabeçalho, larguras, data formatada).
    """
    from pathlib import Path
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    caminho = Path(caminho_xlsx)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"

    cabecalho = ["Data/Hora", "ID", "Item", "Tipo", "Quantidade", "Antes", "Depois"]

    ws.append(cabecalho)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(horizontal="center")

    for col in range(1, len(cabecalho) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    ws.freeze_panes = "A2"

    for m in movimentos:
        delta = float(m.get("delta", 0))
        tipo = "Entrada" if delta > 0 else "Saída"
        qtd = abs(delta)

        ts = m.get("ts", "")
        try:
            ts = datetime.fromisoformat(ts)
        except Exception:
            pass

        ws.append([
            ts,
            m.get("produto_id", ""),
            m.get("nome", ""),
            tipo,
            qtd,
            m.get("estoque_antes", ""),
            m.get("estoque_depois", ""),
        ])

    # Formatos
    for row in ws.iter_rows(min_row=2):
        if hasattr(row[0].value, "year"):
            row[0].number_format = "dd/mm/yyyy hh:mm"

        for idx in (4, 5, 6):
            row[idx].alignment = Alignment(horizontal="center")

    # Ajuste de largura
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(caminho)
    return caminho


