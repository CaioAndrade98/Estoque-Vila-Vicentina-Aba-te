import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import unicodedata
import re
from datetime import datetime, date, timedelta
import csv

from .config import ICONE_ICO  # gui.py e config.py estão em src/
from .estoque_core import (
    listar_produtos,
    criar_produto,
    move_stock_by_id,
    listar_movimentos,
    exportar_movimentos_csv,
    exportar_movimentos_xlsx,
    ProdutoNaoEncontrado,
    EstoqueInsuficiente,
    ProdutoDuplicado,
)


def carregar_produtos() -> list[dict]:
    """Carrega a lista de produtos pelo core (fonte única)."""
    return listar_produtos()


def salvar_produtos(produtos: list[dict]) -> None:
    """Compatibilidade: persistência é feita no core."""
    pass


def gerar_proximo_id(produtos: list[dict]) -> int:
    """Compatibilidade: IDs são gerados no core."""
    raise NotImplementedError("IDs são gerados no core")


def normalizar_nome(nome: str) -> str:
    return " ".join(nome.strip().lower().split())


def remover_acentos(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def normalizar_busca(s: str) -> str:
    """
    Normalização 'premium':
    - lower
    - remove acentos
    - troca pontuação por espaço (coca-cola == coca cola)
    - normaliza múltiplos espaços
    """
    s = remover_acentos(s).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s, flags=re.IGNORECASE)  # pontuação -> espaço
    s = " ".join(s.split())
    return s


def match_prefix_por_palavras(query_tokens: list[str], name_tokens: list[str]) -> bool:
    """query tokens são prefixos dos tokens correspondentes do nome."""
    if not query_tokens:
        return True
    if len(query_tokens) > len(name_tokens):
        return False
    for i, qt in enumerate(query_tokens):
        if not name_tokens[i].startswith(qt):
            return False
    return True


def match_tokens_em_ordem(query_tokens: list[str], name_tokens: list[str]) -> bool:
    """Tokens do query aparecem em ordem (por prefixo), não necessariamente contíguos."""
    if not query_tokens:
        return True
    j = 0
    for nt in name_tokens:
        if nt.startswith(query_tokens[j]):
            j += 1
            if j == len(query_tokens):
                return True
    return False


# =========================
# UX: foco consistente (evita botão "marcado")
# =========================
def _restaurar_foco_menu(root: tk.Tk) -> None:
    """Volta o foco para o menu principal (frame), evitando botão 'marcado'."""
    w = getattr(root, "_menu_focus_widget", None)
    if w is not None:
        try:
            w.focus_set()
        except Exception:
            pass
    else:
        try:
            root.focus_set()
        except Exception:
            pass


def _configurar_fechamento_toplevel(win: tk.Toplevel, root: tk.Tk) -> None:
    """Quando fechar a janela, devolve foco para o menu principal.

    Também dá suporte a "janela única por módulo": se a janela tiver o atributo
    _singleton_key, ela é removida do registro do root ao fechar.
    """
    def _on_close():
        try:
            try:
                # Libera modal, se estiver ativo
                win.grab_release()
            except Exception:
                pass

            win.destroy()
        finally:
            # limpa registro de janela única (se aplicável)
            try:
                key = getattr(win, "_singleton_key", None)
                ow = getattr(root, "_open_windows", None)
                if key and isinstance(ow, dict):
                    if ow.get(key) is win:
                        del ow[key]
            except Exception:
                pass

            root.after(0, lambda: _restaurar_foco_menu(root))

    win.protocol("WM_DELETE_WINDOW", _on_close)


def centralizar_janela(win: tk.Toplevel) -> None:
    """
    Centraliza um Toplevel na tela.
    Funciona mesmo quando a janela está com withdraw(), evitando 'flash' no canto.
    Mantém o tamanho já definido via geometry().
    """
    try:
        win.update_idletasks()
    except Exception:
        pass

    w = h = None
    try:
        g = win.geometry()  # ex: "520x420+0+0"
        m = re.match(r"^(\d+)x(\d+)\+(-?\d+)\+(-?\d+)$", g)
        if m:
            w = int(m.group(1))
            h = int(m.group(2))
    except Exception:
        w = h = None

    if not w or not h:
        # fallback: requested size
        try:
            w = int(win.winfo_reqwidth())
            h = int(win.winfo_reqheight())
        except Exception:
            return

    try:
        sw = int(win.winfo_screenwidth())
        sh = int(win.winfo_screenheight())
        x = (sw - w) // 2
        y = (sh - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")
    except Exception:
        pass



# =========================
# UX: janela única por módulo (evita abrir várias iguais)
# =========================
def _get_open_windows(root: tk.Tk) -> dict:
    ow = getattr(root, "_open_windows", None)
    if not isinstance(ow, dict):
        ow = {}
        setattr(root, "_open_windows", ow)
    return ow


def _singleton_get(root: tk.Tk, key: str) -> tk.Toplevel | None:
    ow = _get_open_windows(root)
    win = ow.get(key)
    try:
        if win is not None and win.winfo_exists():
            # traz para frente
            try:
                win.deiconify()
            except Exception:
                pass
            try:
                win.lift()
                win.focus_force()
            except Exception:
                pass
            return win
    except Exception:
        pass
    # janela morreu; limpa registro
    try:
        if key in ow:
            del ow[key]
    except Exception:
        pass
    return None


def _singleton_register(root: tk.Tk, key: str, win: tk.Toplevel) -> None:
    ow = _get_open_windows(root)
    ow[key] = win
    setattr(win, "_singleton_key", key)

def _parse_iso_ts(ts: str) -> datetime | None:
    try:
        return datetime.fromisoformat(str(ts))
    except Exception:
        return None


def _fmt_dt_br(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y %H:%M")


def _safe_float(x, default=0.0) -> float:
    try:
        return float(x)
    except Exception:
        return float(default)


def abrir_tela_produtos(root: tk.Tk) -> None:
    produtos = carregar_produtos()
    produtos = sorted(produtos, key=lambda p: str(p.get("nome", "")).lower())

    # janela única: se já estiver aberta, traz para frente
    if _singleton_get(root, "estoque") is not None:
        return

    win = tk.Toplevel(root)
    win.title("Estoque")
    win.geometry("720x420")
    _configurar_fechamento_toplevel(win, root)
    _singleton_register(root, "estoque", win)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Itens cadastrados", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

    cols = ("id", "nome", "unidade", "atual", "minimo")
    tree = ttk.Treeview(frame, columns=cols, show="headings", height=14)
    tree.pack(fill="both", expand=True)

    # Ordenação por clique no cabeçalho (profissional)
    sort_state = {"col": None, "desc": False}

    def _coerce(col: str, v: str):
        if col in ("id", "atual", "minimo"):
            try:
                return float(str(v).replace(",", "."))
            except Exception:
                return 0.0
        return str(v).lower()

    def ordenar_por(col: str):
        # alterna asc/desc ao clicar no mesmo cabeçalho
        if sort_state["col"] == col:
            sort_state["desc"] = not sort_state["desc"]
        else:
            sort_state["col"] = col
            sort_state["desc"] = False

        itens = list(tree.get_children(""))
        itens.sort(key=lambda iid: _coerce(col, tree.set(iid, col)), reverse=sort_state["desc"])
        for i, iid in enumerate(itens):
            tree.move(iid, "", i)

    tree.heading("id", text="ID", command=lambda: ordenar_por("id"))
    tree.heading("nome", text="Item", command=lambda: ordenar_por("nome"))
    tree.heading("unidade", text="Unidade", command=lambda: ordenar_por("unidade"))
    tree.heading("atual", text="Atual", command=lambda: ordenar_por("atual"))
    tree.heading("minimo", text="Mínimo", command=lambda: ordenar_por("minimo"))

    tree.column("id", width=60, anchor="center")
    tree.column("nome", width=300)
    tree.column("unidade", width=90, anchor="center")
    tree.column("atual", width=90, anchor="center")
    tree.column("minimo", width=90, anchor="center")

    for p in produtos:
        tree.insert(
            "", "end",
            values=(
                p.get("id", ""),
                p.get("nome", ""),
                p.get("unidade", ""),
                p.get("estoque_atual", 0),
                p.get("estoque_minimo", 0),
            )
        )

    if not produtos:
        messagebox.showinfo("Sem dados", "Nenhum item cadastrado ainda.\nCadastre primeiro pela API ou pelo CLI.")


def abrir_cadastro_produto(root: tk.Tk) -> None:
    # cadastro contínuo (não fecha, sem popup de sucesso)
    win = tk.Toplevel(root)
    win.title("Cadastrar Item")
    win.geometry("520x300")
    win.minsize(520, 300)
    _configurar_fechamento_toplevel(win, root)

    centralizar_janela(win)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Nome do item").pack(anchor="w")
    nome_var = tk.StringVar()
    nome_entry = ttk.Entry(frame, textvariable=nome_var)
    nome_entry.pack(fill="x", pady=(0, 10))

    ttk.Label(frame, text="Unidade (ex: un, kg, L)").pack(anchor="w")
    unidade_var = tk.StringVar()
    unidade_entry = ttk.Entry(frame, textvariable=unidade_var)
    unidade_entry.pack(fill="x", pady=(0, 10))

    ttk.Label(frame, text="Estoque mínimo").pack(anchor="w")
    minimo_var = tk.StringVar()
    minimo_entry = ttk.Entry(frame, textvariable=minimo_var)
    minimo_entry.pack(fill="x", pady=(0, 12))

    # feedback discreto + auto-clear
    status_var = tk.StringVar(value="")
    ttk.Label(frame, textvariable=status_var, foreground="green").pack(anchor="w", pady=(0, 8))
    status_after_id = None

    def _status_ok(msg: str, ms: int = 2000) -> None:
        nonlocal status_after_id
        status_var.set(msg)
        if status_after_id is not None:
            try:
                win.after_cancel(status_after_id)
            except Exception:
                pass
            status_after_id = None
        status_after_id = win.after(ms, lambda: status_var.set(""))

    def salvar():
        nome = nome_var.get().strip()
        unidade = unidade_var.get().strip()
        minimo_txt = minimo_var.get().strip().replace(",", ".")

        if not nome:
            messagebox.showwarning("Atenção", "O nome não pode ficar vazio.")
            nome_entry.focus_set()
            return
        if not unidade:
            messagebox.showwarning("Atenção", "A unidade não pode ficar vazia.")
            unidade_entry.focus_set()
            return
        try:
            minimo = float(minimo_txt)
            if minimo < 0:
                raise ValueError()
        except ValueError:
            messagebox.showwarning("Atenção", "Estoque mínimo inválido. Use um número (ex: 2 ou 2,5).")
            minimo_entry.focus_set()
            return

        try:
            criar_produto(nome=nome, unidade=unidade, estoque_minimo=float(minimo))
        except ProdutoDuplicado:
            messagebox.showerror("Erro", "Já existe um item com esse nome.")
            nome_entry.focus_set()
            return
        except ValueError as e:
            messagebox.showerror("Erro", str(e))
            return

        _status_ok("Item cadastrado com sucesso.", ms=2000)

        nome_var.set("")
        unidade_var.set("")
        minimo_var.set("")
        nome_entry.focus_set()

    ttk.Button(frame, text="Salvar", command=salvar, takefocus=False).pack(anchor="e")

    nome_entry.bind("<Return>", lambda e: unidade_entry.focus_set())
    unidade_entry.bind("<Return>", lambda e: minimo_entry.focus_set())
    minimo_entry.bind("<Return>", lambda e: salvar())

    nome_entry.focus_set()


def atualizar_estoque(produto_id: int, delta: float, motivo: str | None = None) -> None:
    """Movimenta estoque usando o core (regra única)."""
    try:
        move_stock_by_id(produto_id, delta, motivo=motivo)
    except ProdutoNaoEncontrado as e:
        raise ValueError(str(e))
    except EstoqueInsuficiente as e:
        raise ValueError(str(e))


def abrir_movimento(root: tk.Tk, tipo: str) -> None:
    # tipo: "entrada" ou "saida"
    produtos = carregar_produtos()
    produtos = sorted(produtos, key=lambda p: str(p.get("nome", "")).lower())

    if not produtos:
        messagebox.showinfo("Sem itens", "Cadastre um item antes.")
        return

    # Mapa id->produto para prévia (estoque atual/mínimo/unidade)
    id_para_produto: dict[int, dict] = {}
    for p in produtos:
        try:
            pid = int(p.get("id", 0))
        except Exception:
            continue
        id_para_produto[pid] = p

    win = tk.Toplevel(root)
    win.title("Entrada de Estoque" if tipo == "entrada" else "Saída de Estoque")
    win.geometry("560x560")  # ligeiro aumento para caber prévia/status
    centralizar_janela(win)
    _configurar_fechamento_toplevel(win, root)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    # ---------- BUSCA ----------
    ttk.Label(frame, text="Buscar item").pack(anchor="w")

    # Exibição SEM ID: só nome; se repetir, inclui unidade; se ainda repetir, (2), (3)...
    itens_exibicao: list[str] = []
    texto_para_id: dict[str, int] = {}

    # Índices para busca premium (somente pelo NOME)
    item_para_nome_norm: dict[str, str] = {}
    item_para_tokens: dict[str, list[str]] = {}

    usados: dict[str, int] = {}

    for p in produtos:
        nome = str(p.get("nome", "")).strip()
        unidade = str(p.get("unidade", "")).strip()
        pid = int(p.get("id", 0))

        base = nome
        if base in usados:
            base = f"{nome} ({unidade})" if unidade else nome

        n = usados.get(base, 0) + 1
        usados[base] = n
        texto_exibido = base if n == 1 else f"{base} ({n})"

        itens_exibicao.append(texto_exibido)
        texto_para_id[texto_exibido] = pid

        nn = normalizar_busca(nome)
        item_para_nome_norm[texto_exibido] = nn
        item_para_tokens[texto_exibido] = nn.split() if nn else []

    itens_originais = itens_exibicao.copy()

    busca_var = tk.StringVar(value="")
    ent_busca = ttk.Entry(frame, textvariable=busca_var)
    ent_busca.pack(fill="x", pady=(0, 6))

    # Status: contador e vazio
    status_frame = ttk.Frame(frame)
    status_frame.pack(fill="x", pady=(0, 4))

    resultados_var = tk.StringVar(value=f"{len(itens_originais)} itens")
    ttk.Label(status_frame, textvariable=resultados_var).pack(side="left")

    vazio_var = tk.StringVar(value="")
    ttk.Label(status_frame, textvariable=vazio_var).pack(side="right")

    # Feedback padrão (3 níveis) + auto-clear
    status_var = tk.StringVar(value="")
    lbl_status = ttk.Label(frame, textvariable=status_var, foreground="green")
    lbl_status.pack(anchor="w", pady=(0, 6))

    status_after_id = None
    CORES = {
        "success": "green",
        "warn": "#b26a00",
        "error": "#b00020",
        "info": "#1a1a1a",
    }

    def _status(msg: str, kind: str = "info", ms: int = 2200) -> None:
        nonlocal status_after_id
        lbl_status.configure(foreground=CORES.get(kind, CORES["info"]))
        status_var.set(msg)

        if status_after_id is not None:
            try:
                win.after_cancel(status_after_id)
            except Exception:
                pass
            status_after_id = None

        status_after_id = win.after(ms, lambda: status_var.set(""))

    ttk.Label(frame, text="Selecione um item").pack(anchor="w")

    # ---------- LISTBOX + SCROLLBAR ----------
    list_frame = ttk.Frame(frame)
    list_frame.pack(fill="both", expand=True, pady=(0, 10))

    yscroll = ttk.Scrollbar(list_frame, orient="vertical")
    yscroll.pack(side="right", fill="y")

    lb = tk.Listbox(
        list_frame,
        height=10,
        activestyle="none",
        exportselection=False,
        yscrollcommand=yscroll.set,

        bg="#ffffff",
        fg="#1f2933",
        selectbackground="#2563eb",
        selectforeground="#ffffff",
        highlightthickness=1,
        highlightbackground="#d1d5db",
        relief="flat",
        borderwidth=0,
    )
    lb.pack(side="left", fill="both", expand=True)
    yscroll.config(command=lb.yview)

    for item in itens_originais:
        lb.insert("end", item)

    # ---------- PRÉVIA DO ESTOQUE DO ITEM SELECIONADO ----------
    preview_var = tk.StringVar(value="Selecione um item para ver o estoque atual.")
    lbl_preview = ttk.Label(frame, textvariable=preview_var)
    lbl_preview.pack(anchor="w", pady=(0, 10))

    def _atualizar_status(qtd: int):
        resultados_var.set(f"{qtd} item" if qtd == 1 else f"{qtd} itens")
        vazio_var.set("Nenhum item encontrado" if qtd == 0 else "")

    def _filtrar_rankeado_premium(texto_busca: str) -> list[str]:
        q = normalizar_busca(texto_busca)
        if not q:
            return itens_originais

        qtoks = q.split()

        grupo_a: list[str] = []
        grupo_b: list[str] = []
        grupo_c: list[str] = []
        grupo_d: list[str] = []

        for item_txt in itens_originais:
            nome_norm = item_para_nome_norm.get(item_txt, "")
            ntoks = item_para_tokens.get(item_txt, [])

            if nome_norm.startswith(q):
                grupo_a.append(item_txt)
            elif match_prefix_por_palavras(qtoks, ntoks):
                grupo_b.append(item_txt)
            elif q in nome_norm:
                grupo_c.append(item_txt)
            elif match_tokens_em_ordem(qtoks, ntoks):
                grupo_d.append(item_txt)

        return grupo_a + grupo_b + grupo_c + grupo_d

    def _render_lista(itens: list[str]) -> None:
        lb.delete(0, "end")
        for it in itens:
            lb.insert("end", it)

    _after_id = None

    def obter_item_selecionado() -> str | None:
        sel = lb.curselection()
        if not sel:
            return None
        idx = int(sel[0])
        try:
            return lb.get(idx)
        except Exception:
            return None

    def _atualizar_preview():
        item_txt = obter_item_selecionado()
        if not item_txt:
            preview_var.set("Selecione um item para ver o estoque atual.")
            return

        pid = texto_para_id.get(item_txt)
        if not pid:
            preview_var.set("Não foi possível identificar o item selecionado.")
            return

        p = id_para_produto.get(int(pid), {})
        unidade = str(p.get("unidade", "")).strip()
        atual = p.get("estoque_atual", 0)
        minimo = p.get("estoque_minimo", 0)

        sufixo_un = f" {unidade}" if unidade else ""
        preview_var.set(f"Estoque atual: {atual}{sufixo_un}  |  Mínimo: {minimo}{sufixo_un}")

    def aplicar_filtro():
        filtradas = _filtrar_rankeado_premium(busca_var.get())
        _render_lista(filtradas)
        _atualizar_status(len(filtradas))

        if len(filtradas) == 1:
            lb.selection_clear(0, "end")
            lb.selection_set(0)
            lb.activate(0)
            lb.see(0)
            _atualizar_preview()
            _avaliar_warn_saida()
        else:
            preview_var.set("Selecione um item para ver o estoque atual.")

    def on_busca_keyrelease(event=None):
        nonlocal _after_id
        if _after_id is not None:
            win.after_cancel(_after_id)
        _after_id = win.after(120, aplicar_filtro)

    def selecionar_primeiro_da_lista():
        if lb.size() > 0:
            lb.selection_clear(0, "end")
            lb.selection_set(0)
            lb.activate(0)
            lb.see(0)
            _atualizar_preview()
            _avaliar_warn_saida()

    def ir_para_lista(event=None):
        if lb.size() > 0:
            selecionar_primeiro_da_lista()
            lb.focus_set()
        return "break"

    def on_busca_enter(event=None):
        if lb.size() > 0:
            selecionar_primeiro_da_lista()
            ent_qtd.focus_set()
        else:
            _status("Nenhum item para selecionar.", kind="warn", ms=2200)
        return "break"

    def on_busca_escape(event=None):
        busca_var.set("")
        _render_lista(itens_originais)
        _atualizar_status(len(itens_originais))
        lb.selection_clear(0, "end")
        preview_var.set("Selecione um item para ver o estoque atual.")
        ent_busca.focus_set()
        return "break"

    def on_lista_enter(event=None):
        ent_qtd.focus_set()
        return "break"

    def on_lista_double(event=None):
        ent_qtd.focus_set()
        return "break"

    def on_lista_escape(event=None):
        ent_busca.focus_set()
        return "break"

    ent_busca.bind("<KeyRelease>", on_busca_keyrelease)
    ent_busca.bind("<Down>", ir_para_lista)
    ent_busca.bind("<Return>", on_busca_enter)
    ent_busca.bind("<Escape>", on_busca_escape)

    ent_busca.focus_set()
    _atualizar_status(len(itens_originais))

    # ---------- QUANTIDADE ----------
    ttk.Label(frame, text="Quantidade (ex: 1,5)").pack(anchor="w")
    qtd_var = tk.StringVar()
    ent_qtd = ttk.Entry(frame, textvariable=qtd_var)
    ent_qtd.pack(fill="x", pady=(0, 12))

    ttk.Label(frame, text="Motivo (opcional)").pack(anchor="w")
    motivo_var = tk.StringVar()
    ent_motivo = ttk.Entry(frame, textvariable=motivo_var)
    ent_motivo.pack(fill="x", pady=(0, 12))

    
    # ✅ Aviso imediato: SAÍDA com quantidade maior que estoque atual
    _warn_ativa = False

    def _avaliar_warn_saida(event=None):
        nonlocal _warn_ativa, status_after_id

        try:
            btn_confirmar.state(["!disabled"])
        except Exception:
            pass

        if tipo != "saida":
            _warn_ativa = False
            return

        item_txt = obter_item_selecionado()
        if not item_txt:
            _warn_ativa = False
            return

        pid = texto_para_id.get(item_txt)
        if not pid:
            _warn_ativa = False
            return

        p = id_para_produto.get(int(pid), {})
        atual = _safe_float(p.get("estoque_atual", 0), 0.0)

        txt = qtd_var.get().strip()
        if not txt:
            if _warn_ativa:
                status_var.set("")
            _warn_ativa = False
            return

        try:
            qtd = float(txt.replace(",", "."))
        except Exception:
            if _warn_ativa:
                status_var.set("")
            _warn_ativa = False
            return

        if qtd > atual:
            _warn_ativa = True

            if status_after_id is not None:
                try:
                    win.after_cancel(status_after_id)
                except Exception:
                    pass
                status_after_id = None

            lbl_status.configure(foreground=CORES.get("warn", "#b26a00"))
            status_var.set(f"⚠ Atenção: {qtd} é maior que o estoque atual ({atual}).")

            try:
                btn_confirmar.state(["disabled"])
            except Exception:
                pass
        else:
            if _warn_ativa:
                status_var.set("")
            _warn_ativa = False
            try:
                btn_confirmar.state(["!disabled"])
            except Exception:
                pass

    def on_lista_select(event=None):
        _atualizar_preview()
        _avaliar_warn_saida()

    lb.bind("<<ListboxSelect>>", on_lista_select)
    lb.bind("<Return>", on_lista_enter)
    lb.bind("<Double-Button-1>", on_lista_double)
    lb.bind("<Escape>", on_lista_escape)

    ent_qtd.bind("<KeyRelease>", _avaliar_warn_saida)

    def confirmar(foco_busca: bool = False):
        try:
            if "disabled" in btn_confirmar.state():
                lbl_status.configure(foreground=CORES.get("warn", "#b26a00"))
                status_var.set("⚠ Corrija a quantidade (está maior que o estoque atual).")
                ent_qtd.focus_set()
                return
        except Exception:
            pass

        item_txt = obter_item_selecionado()
        if not item_txt:
            _status("Selecione um item na lista.", kind="warn", ms=2500)
            lb.focus_set()
            return

        qtd_txt = qtd_var.get().strip().replace(",", ".")
        try:
            qtd = float(qtd_txt)
            if qtd <= 0:
                raise ValueError()
        except ValueError:
            _status("Quantidade inválida. Use um número maior que 0 (ex: 1 ou 1,5).", kind="warn", ms=2800)
            ent_qtd.focus_set()
            return

        pid = texto_para_id.get(item_txt)
        if not pid:
            messagebox.showerror("Erro", "Não foi possível identificar o item selecionado.")
            return

        try:
            delta = qtd if tipo == "entrada" else -qtd
            motivo = motivo_var.get().strip() or None
            atualizar_estoque(int(pid), float(delta), motivo=motivo)

        except ValueError as e:
            messagebox.showerror("Erro", str(e))
            ent_qtd.focus_set()
            return

        _status(
            "Entrada registrada com sucesso." if tipo == "entrada" else "Saída registrada com sucesso.",
            kind="success",
            ms=2000,
        )

        qtd_var.set("")
        motivo_var.set("")

        novos = carregar_produtos()
        for pp in novos:
            try:
                pid2 = int(pp.get("id", 0))
            except Exception:
                continue
            id_para_produto[pid2] = pp

        _atualizar_preview()
        _avaliar_warn_saida()

        if foco_busca:
            ent_busca.focus_set()
        else:
            ent_qtd.focus_set()

    btn_confirmar = ttk.Button(frame, text="Confirmar", command=lambda: confirmar(False), takefocus=False)
    btn_confirmar.pack(anchor="e")

    ent_qtd.bind("<Return>", lambda e: confirmar(False))
    win.bind("<Control-Return>", lambda e: (confirmar(True), "break"))

    def _esc_qtd(event=None):
        qtd_var.set("")
        ent_busca.focus_set()
        _avaliar_warn_saida()
        return "break"

    ent_qtd.bind("<Escape>", _esc_qtd)


def abrir_abaixo_minimo(root: tk.Tk) -> None:
    produtos = carregar_produtos()
    abaixo = [
        p for p in produtos
        if float(p.get("estoque_atual", 0)) < float(p.get("estoque_minimo", 0))
    ]
    abaixo = sorted(abaixo, key=lambda p: str(p.get("nome", "")).lower())

    win = tk.Toplevel(root)
    win.title("Abaixo do mínimo")
    win.geometry("720x420")
    _configurar_fechamento_toplevel(win, root)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Itens abaixo do estoque mínimo", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

    status_var = tk.StringVar(value="")
    lbl = ttk.Label(frame, textvariable=status_var, foreground="green")
    lbl.pack(anchor="w", pady=(0, 8))
    status_after_id = None

    def _status_ok(msg: str, ms: int = 2500) -> None:
        nonlocal status_after_id
        status_var.set(msg)
        if status_after_id is not None:
            try:
                win.after_cancel(status_after_id)
            except Exception:
                pass
            status_after_id = None
        status_after_id = win.after(ms, lambda: status_var.set(""))

    cols = ("nome", "unidade", "atual", "minimo")
    tree = ttk.Treeview(frame, columns=cols, show="headings", height=14)
    tree.pack(fill="both", expand=True)

    tree.heading("nome", text="Item")
    tree.heading("unidade", text="Unidade")
    tree.heading("atual", text="Atual")
    tree.heading("minimo", text="Mínimo")

    tree.column("nome", width=360)
    tree.column("unidade", width=90, anchor="center")
    tree.column("atual", width=90, anchor="center")
    tree.column("minimo", width=90, anchor="center")

    for p in abaixo:
        tree.insert(
            "", "end",
            values=(
                p.get("nome", ""),
                p.get("unidade", ""),
                p.get("estoque_atual", 0),
                p.get("estoque_minimo", 0),
            )
        )

    if not abaixo:
        _status_ok("Nenhum item está abaixo do mínimo.", ms=2500)


def abrir_historico(root: tk.Tk) -> None:
    LIMITE_PADRAO = 300
    PAGE_SIZE = 800  # incremental render para evitar lag no Treeview

    # janela única
    if _singleton_get(root, "historico") is not None:
        return

    win = tk.Toplevel(root)
    win.title("Histórico de movimentações")
    _configurar_fechamento_toplevel(win, root)
    _singleton_register(root, "historico", win)

    win.state("zoomed")  # tela cheia no Windows
    win.minsize(1020, 720)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Histórico de movimentações", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

    filtros = ttk.Frame(frame)
    filtros.pack(fill="x", pady=(0, 10))

    ttk.Label(filtros, text="Filtrar por item").pack(side="left")
    filtro_var = tk.StringVar(value="")
    ent_filtro = ttk.Entry(filtros, textvariable=filtro_var, width=30)
    ent_filtro.pack(side="left", padx=(8, 16))

    ttk.Label(filtros, text="Limite").pack(side="left")
    limite_var = tk.StringVar(value=str(LIMITE_PADRAO))
    ent_limite = ttk.Entry(filtros, textvariable=limite_var, width=8)
    ent_limite.pack(side="left", padx=(8, 16))

    ok_var = tk.StringVar(value="")
    ttk.Label(filtros, textvariable=ok_var, foreground="green").pack(side="right")

    ok_after_id = None

    def _ok(msg: str, ms: int = 2500) -> None:
        nonlocal ok_after_id
        ok_var.set(msg)
        if ok_after_id is not None:
            try:
                win.after_cancel(ok_after_id)
            except Exception:
                pass
            ok_after_id = None
        ok_after_id = win.after(ms, lambda: ok_var.set(""))

    status_var = tk.StringVar(value="")
    ttk.Label(filtros, textvariable=status_var).pack(side="right", padx=(0, 10))

    cols = ("ts", "nome", "tipo", "motivo", "qtd", "antes", "depois")

    table_frame = ttk.Frame(frame)
    table_frame.pack(fill="both", expand=True)

    tree_frame = ttk.Frame(table_frame)
    tree_frame.pack(fill="both", expand=True)

    tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=16)
    tree.pack(side="left", fill="both", expand=True)

    yscroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    yscroll.pack(side="right", fill="y")

    xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
    xscroll.pack(side="bottom", fill="x")

    tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

    tree.heading("ts", text="Data/Hora")
    tree.heading("nome", text="Item")
    tree.heading("tipo", text="Tipo")
    tree.heading("motivo", text="Motivo")
    tree.heading("qtd", text="Qtd")
    tree.heading("antes", text="Antes")
    tree.heading("depois", text="Depois")

    tree.column("ts", width=150, anchor="center")
    tree.column("nome", width=300)
    tree.column("tipo", width=80, anchor="center")
    tree.column("motivo", width=220)
    tree.column("qtd", width=80, anchor="center")
    tree.column("antes", width=80, anchor="center")
    tree.column("depois", width=80, anchor="center")

    for col in cols:
        tree.heading(col, anchor="center")

    movimentos_cache: list[dict] = []
    filtrados_cache: list[dict] = []
    render_limit = PAGE_SIZE

    def _parse_limite() -> int:
        try:
            n = int(limite_var.get().strip())
        except Exception:
            n = LIMITE_PADRAO
        if n < 1:
            n = 1
        if n > 5000:
            n = 5000
        return n

    def _format_tipo(delta: float) -> str:
        return "Entrada" if float(delta) > 0 else "Saída"

    def _format_qtd(delta: float) -> float:
        return abs(float(delta))

    def _recalcular_filtrados() -> None:
        nonlocal filtrados_cache
        termo = normalizar_busca(filtro_var.get().strip())

        filtrados: list[dict] = []
        for m in movimentos_cache:
            nome = str(m.get("nome", ""))
            nome_norm = normalizar_busca(nome)
            if termo and termo not in nome_norm:
                continue
            filtrados.append(m)

        filtrados_cache = filtrados

    def _render() -> None:
        tree.delete(*tree.get_children())

        total = len(filtrados_cache)
        exibidos = filtrados_cache[:render_limit]

        for m in exibidos:
            nome = str(m.get("nome", ""))
            delta = _safe_float(m.get("delta", 0), 0.0)
            dt = _parse_iso_ts(m.get("ts", ""))
            ts = _fmt_dt_br(dt)
            motivo = str(m.get("motivo", "") or "")

            tree.insert(
                "", "end",
                values=(
                    ts,
                    nome,
                    _format_tipo(delta),
                    motivo,
                    _format_qtd(delta),
                    m.get("estoque_antes", ""),
                    m.get("estoque_depois", ""),
                ),
            )

        if total > render_limit:
            status_var.set(f"{total} registro(s) — exibindo {render_limit}.")
        else:
            status_var.set(f"{total} registro(s)")

        # botão mais: habilita só se ainda houver mais pra carregar
        try:
            if total > render_limit:
                btn_mais.state(["!disabled"])
            else:
                btn_mais.state(["disabled"])
        except Exception:
            pass

    def carregar() -> None:
        nonlocal movimentos_cache, render_limit
        limite = _parse_limite()
        movimentos = listar_movimentos(limite=limite)
        movimentos_cache = list(reversed(movimentos))  # mais recentes no topo
        render_limit = PAGE_SIZE
        _recalcular_filtrados()
        _render()

    def aplicar_filtro() -> None:
        nonlocal render_limit
        render_limit = PAGE_SIZE
        _recalcular_filtrados()
        _render()

    def carregar_mais() -> None:
        nonlocal render_limit
        if render_limit < len(filtrados_cache):
            render_limit = min(render_limit + PAGE_SIZE, len(filtrados_cache))
            _render()

    def limpar_filtro() -> None:
        filtro_var.set("")
        aplicar_filtro()

    def exportar_csv() -> None:
        nome_sugerido = f"historico_estoque_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        caminho = filedialog.asksaveasfilename(
            title="Salvar histórico em CSV",
            defaultextension=".csv",
            initialfile=nome_sugerido,
            filetypes=[("CSV", "*.csv")],
        )
        if not caminho:
            return

        try:
            termo = normalizar_busca(filtro_var.get().strip())
            linhas = []
            for m in movimentos_cache:
                nome = str(m.get("nome", ""))
                nome_norm = normalizar_busca(nome)
                if termo and termo not in nome_norm:
                    continue
                linhas.append(m)

            with open(caminho, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["data_hora", "item", "tipo", "motivo", "qtd", "antes", "depois"])
                for m in linhas:
                    delta = _safe_float(m.get("delta", 0), 0.0)
                    dt = _parse_iso_ts(m.get("ts", ""))
                    w.writerow([
                        _fmt_dt_br(dt),
                        str(m.get("nome", "")),
                        _format_tipo(delta),
                        str(m.get("motivo", "") or ""),
                        _format_qtd(delta),
                        m.get("estoque_antes", ""),
                        m.get("estoque_depois", ""),
                    ])

            _ok("CSV exportado com sucesso.", ms=2500)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível exportar o CSV.\n\n{e}")

    def exportar_excel() -> None:
        nome_sugerido = f"historico_estoque_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        caminho = filedialog.asksaveasfilename(
            title="Salvar histórico em Excel",
            defaultextension=".xlsx",
            initialfile=nome_sugerido,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not caminho:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            termo = normalizar_busca(filtro_var.get().strip())
            linhas = []
            for m in movimentos_cache:
                nome = str(m.get("nome", ""))
                nome_norm = normalizar_busca(nome)
                if termo and termo not in nome_norm:
                    continue
                linhas.append(m)

            wb = Workbook()
            ws = wb.active
            ws.title = "Histórico"

            header = ["Data/Hora", "Item", "Tipo", "Motivo", "Qtd", "Antes", "Depois"]
            ws.append(header)

            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="DDDDDD")
            header_align = Alignment(horizontal="center")

            for col in range(1, len(header) + 1):
                c = ws.cell(row=1, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align

            for m in linhas:
                delta = _safe_float(m.get("delta", 0), 0.0)
                dt = _parse_iso_ts(m.get("ts", ""))
                ws.append([
                    _fmt_dt_br(dt),
                    str(m.get("nome", "")),
                    _format_tipo(delta),
                    str(m.get("motivo", "") or ""),
                    _format_qtd(delta),
                    m.get("estoque_antes", ""),
                    m.get("estoque_depois", ""),
                ])

            # largura
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

            wb.save(caminho)
            _ok("Excel exportado com sucesso.", ms=2500)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível exportar o Excel.\n\n{e}")

    botoes = ttk.Frame(frame)
    botoes.pack(fill="x", pady=(10, 0))

    # 'Carregar mais' (incremental) para evitar lag com muitos registros
    btn_mais = ttk.Button(botoes, text=f"Carregar mais {PAGE_SIZE}", command=carregar_mais, takefocus=False)
    btn_mais.pack(side="left")

    ttk.Button(botoes, text="Atualizar", command=carregar, takefocus=False).pack(side="right")
    ttk.Button(botoes, text="Exportar Excel", command=exportar_excel, takefocus=False).pack(side="right", padx=(0, 8))
    ttk.Button(botoes, text="Exportar CSV", command=exportar_csv, takefocus=False).pack(side="right", padx=(0, 8))
    ttk.Button(botoes, text="Limpar filtro", command=limpar_filtro, takefocus=False).pack(side="right", padx=(0, 8))

    ent_filtro.bind("<Return>", lambda e: aplicar_filtro())
    ent_filtro.bind("<Escape>", lambda e: limpar_filtro())

    carregar()
    ent_filtro.focus_set()


# ============================================================
# 1) DASHBOARD - VISÃO GERAL
# ============================================================
def abrir_dashboard(root: tk.Tk) -> None:
    # janela única
    if _singleton_get(root, "dashboard") is not None:
        return

    win = tk.Toplevel(root)
    win.title("Visão Geral")
    win.state("zoomed")   # tela cheia no Windows
    win.minsize(1180, 640)
    _configurar_fechamento_toplevel(win, root)
    _singleton_register(root, "dashboard", win)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Visão Geral", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 10))

    # Cards simples (sem refatorar, só frames)
    top = ttk.Frame(frame)
    top.pack(fill="x")

    card_a = ttk.Frame(top, padding=12)
    card_b = ttk.Frame(top, padding=12)
    card_c = ttk.Frame(top, padding=12)
    card_a.pack(side="left", fill="x", expand=True, padx=(0, 8))
    card_b.pack(side="left", fill="x", expand=True, padx=(0, 8))
    card_c.pack(side="left", fill="x", expand=True)

    total_var = tk.StringVar(value="0")
    abaixo_var = tk.StringVar(value="0")
    alertas_var = tk.StringVar(value="0")

    ttk.Label(card_a, text="Itens cadastrados", font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(card_a, textvariable=total_var, font=("Segoe UI", 20, "bold")).pack(anchor="w")

    ttk.Label(card_b, text="Abaixo do mínimo", font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(card_b, textvariable=abaixo_var, font=("Segoe UI", 20, "bold")).pack(anchor="w")

    ttk.Label(card_c, text="Alertas (zerados / quase)", font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(card_c, textvariable=alertas_var, font=("Segoe UI", 20, "bold")).pack(anchor="w")

    mid = ttk.Frame(frame)
    mid.pack(fill="both", expand=True, pady=(12, 0))

    # Top 5 cobertura
    left = ttk.Frame(mid)
    left.pack(side="left", fill="both", expand=True, padx=(0, 8))

    ttk.Label(left, text="Top 5 menor cobertura (atual / mínimo)", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

    cols_cov = ("nome", "un", "atual", "min", "cob")
    tree_cov = ttk.Treeview(left, columns=cols_cov, show="headings", height=10)
    tree_cov.pack(side="top", fill="both", expand=True)

    xscroll_cov = ttk.Scrollbar(left, orient="horizontal", command=tree_cov.xview)
    xscroll_cov.pack(side="bottom", fill="x")
    tree_cov.configure(xscrollcommand=xscroll_cov.set)

    # Ordenação por clique no cabeçalho (profissional)
    _sort_cov = {"col": None, "desc": False}

    def _coerce_cov(col: str, v: str):
        if col in ("atual", "min"):
            try:
                return float(str(v).replace(",", "."))
            except Exception:
                return 0.0
        if col == "cob":
            # ex: "0.85x"
            try:
                return float(str(v).lower().replace("x", "").replace(",", ".").strip())
            except Exception:
                return 0.0
        return str(v).lower()

    def ordenar_cov(col: str):
        if _sort_cov["col"] == col:
            _sort_cov["desc"] = not _sort_cov["desc"]
        else:
            _sort_cov["col"] = col
            _sort_cov["desc"] = False

        itens = list(tree_cov.get_children(""))
        itens.sort(key=lambda iid: _coerce_cov(col, tree_cov.set(iid, col)), reverse=_sort_cov["desc"])
        for i, iid in enumerate(itens):
            tree_cov.move(iid, "", i)

    tree_cov.heading("nome", text="Item", command=lambda: ordenar_cov("nome"))
    tree_cov.heading("un", text="Unid.", command=lambda: ordenar_cov("un"))
    tree_cov.heading("atual", text="Atual", command=lambda: ordenar_cov("atual"))
    tree_cov.heading("min", text="Mínimo", command=lambda: ordenar_cov("min"))
    tree_cov.heading("cob", text="Cobertura", command=lambda: ordenar_cov("cob"))

    tree_cov.column("nome", width=320)
    tree_cov.column("un", width=70, anchor="center")
    tree_cov.column("atual", width=80, anchor="center")
    tree_cov.column("min", width=80, anchor="center")
    tree_cov.column("cob", width=90, anchor="center")

    # Últimas 10 movimentações
    right = ttk.Frame(mid)
    right.pack(side="left", fill="both", expand=True)

    ttk.Label(right, text="Últimas 10 movimentações", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

    cols_mov = ("ts", "nome", "tipo", "qtd")
    tree_mov = ttk.Treeview(right, columns=cols_mov, show="headings", height=10)
    tree_mov.pack(side="top", fill="both", expand=True)

    xscroll_mov = ttk.Scrollbar(right, orient="horizontal", command=tree_mov.xview)
    xscroll_mov.pack(side="bottom", fill="x")
    tree_mov.configure(xscrollcommand=xscroll_mov.set)

    # Ordenação por clique no cabeçalho (profissional)
    _sort_mov = {"col": None, "desc": False}

    def _coerce_mov(col: str, v: str):
        if col == "qtd":
            try:
                return float(str(v).replace(",", "."))
            except Exception:
                return 0.0
        if col == "ts":
            # formato BR: dd/mm/YYYY HH:MM
            try:
                return datetime.strptime(str(v), "%d/%m/%Y %H:%M")
            except Exception:
                return datetime.min
        return str(v).lower()

    def ordenar_mov(col: str):
        if _sort_mov["col"] == col:
            _sort_mov["desc"] = not _sort_mov["desc"]
        else:
            _sort_mov["col"] = col
            _sort_mov["desc"] = False

        itens = list(tree_mov.get_children(""))
        itens.sort(key=lambda iid: _coerce_mov(col, tree_mov.set(iid, col)), reverse=_sort_mov["desc"])
        for i, iid in enumerate(itens):
            tree_mov.move(iid, "", i)

    tree_mov.heading("ts", text="Data/Hora", command=lambda: ordenar_mov("ts"))
    tree_mov.heading("nome", text="Item", command=lambda: ordenar_mov("nome"))
    tree_mov.heading("tipo", text="Tipo", command=lambda: ordenar_mov("tipo"))
    tree_mov.heading("qtd", text="Qtd", command=lambda: ordenar_mov("qtd"))

    tree_mov.column("ts", width=140, anchor="center")
    tree_mov.column("nome", width=320)
    tree_mov.column("tipo", width=70, anchor="center")
    tree_mov.column("qtd", width=70, anchor="center")

    status_var = tk.StringVar(value="")
    ttk.Label(frame, textvariable=status_var).pack(anchor="w", pady=(10, 0))

    def atualizar():
        produtos = carregar_produtos()
        total = len(produtos)

        abaixo = 0
        alertas = 0

        cov_list = []
        for p in produtos:
            atual = _safe_float(p.get("estoque_atual", 0), 0.0)
            minimo = _safe_float(p.get("estoque_minimo", 0), 0.0)
            if atual < minimo:
                abaixo += 1

            # alertas: zerado ou "quase" (<= minimo)
            if atual <= 0:
                alertas += 1
            elif minimo > 0 and atual <= minimo:
                alertas += 1

            # cobertura
            if minimo <= 0:
                cobertura = float("inf")  # sem mínimo definido, não entra no top ruim
            else:
                cobertura = atual / minimo

            cov_list.append((cobertura, p))

        total_var.set(str(total))
        abaixo_var.set(str(abaixo))
        alertas_var.set(str(alertas))

        # Top 5 menor cobertura (ignorando inf)
        tree_cov.delete(*tree_cov.get_children())
        cov_list = [x for x in cov_list if x[0] != float("inf")]
        cov_list.sort(key=lambda x: x[0])
        for cobertura, p in cov_list[:5]:
            un = str(p.get("unidade", "")).strip()
            atual = _safe_float(p.get("estoque_atual", 0), 0.0)
            minimo = _safe_float(p.get("estoque_minimo", 0), 0.0)
            tree_cov.insert(
                "", "end",
                values=(p.get("nome", ""), un, atual, minimo, f"{cobertura:.2f}x"),
            )

        # Últimas 10 movimentações
        tree_mov.delete(*tree_mov.get_children())
        movs = listar_movimentos(limite=10)
        movs = list(reversed(movs))  # mais recentes primeiro
        for m in movs:
            dt = _parse_iso_ts(m.get("ts", ""))
            nome = str(m.get("nome", ""))
            delta = _safe_float(m.get("delta", 0), 0.0)
            tipo = "Entrada" if delta > 0 else "Saída"
            qtd = abs(delta)
            tree_mov.insert("", "end", values=(_fmt_dt_br(dt), nome, tipo, qtd))

        status_var.set(f"Atualizado em {_fmt_dt_br(datetime.now())}")

    botoes = ttk.Frame(frame)
    botoes.pack(fill="x", pady=(10, 0))
    ttk.Button(botoes, text="Atualizar", command=atualizar, takefocus=False).pack(side="right")

    atualizar()


# ============================================================
# 2) RELATÓRIOS POR PERÍODO
# ============================================================
def abrir_relatorios_periodo(root: tk.Tk) -> None:
    # janela única
    if _singleton_get(root, "relatorios_periodo") is not None:
        return

    win = tk.Toplevel(root)
    win.title("Relatórios por período")
    win.state("zoomed")  # tela cheia no Windows
    win.minsize(1150, 720)
    _configurar_fechamento_toplevel(win, root)
    _singleton_register(root, "relatorios_periodo", win)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Relatórios por período", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 10))

    filtros = ttk.Frame(frame)
    filtros.pack(fill="x", pady=(0, 10))

    ttk.Label(filtros, text="De (YYYY-MM-DD)").pack(side="left")
    de_var = tk.StringVar(value="")
    ent_de = ttk.Entry(filtros, textvariable=de_var, width=12)
    ent_de.pack(side="left", padx=(6, 12))

    ttk.Label(filtros, text="Até (YYYY-MM-DD)").pack(side="left")
    ate_var = tk.StringVar(value="")
    ent_ate = ttk.Entry(filtros, textvariable=ate_var, width=12)
    ent_ate.pack(side="left", padx=(6, 12))

    # atalhos rápidos
    def _set_periodo(dias: int):
        hoje = date.today()
        ini = hoje - timedelta(days=dias)
        de_var.set(ini.isoformat())
        ate_var.set(hoje.isoformat())

    ttk.Button(filtros, text="7 dias", command=lambda: _set_periodo(7), takefocus=False).pack(side="left", padx=(0, 6))
    ttk.Button(filtros, text="30 dias", command=lambda: _set_periodo(30), takefocus=False).pack(side="left", padx=(0, 6))
    ttk.Button(filtros, text="90 dias", command=lambda: _set_periodo(90), takefocus=False).pack(side="left", padx=(0, 12))

    status_var = tk.StringVar(value="")
    ttk.Label(filtros, textvariable=status_var).pack(side="right")

    # tabelas
    mid = ttk.Frame(frame)
    mid.pack(fill="both", expand=True)

    left = ttk.Frame(mid)
    left.pack(side="left", fill="both", expand=True, padx=(0, 8))

    ttk.Label(left, text="Totais por item (Entradas / Saídas)", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
    cols = ("nome", "entradas", "saidas", "saldo")
    tree = ttk.Treeview(left, columns=cols, show="headings", height=18)
    tree.pack(fill="both", expand=True)

    tree.heading("nome", text="Item")
    tree.heading("entradas", text="Entradas")
    tree.heading("saidas", text="Saídas")
    tree.heading("saldo", text="Saldo")

    tree.column("nome", width=320)
    tree.column("entradas", width=90, anchor="center")
    tree.column("saidas", width=90, anchor="center")
    tree.column("saldo", width=90, anchor="center")

    right = ttk.Frame(mid)
    right.pack(side="left", fill="both", expand=True)

    ttk.Label(right, text="Itens mais movimentados (volume)", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
    cols2 = ("nome", "volume")
    tree2 = ttk.Treeview(right, columns=cols2, show="headings", height=18)
    tree2.pack(fill="both", expand=True)

    tree2.heading("nome", text="Item")
    tree2.heading("volume", text="Volume")

    tree2.column("nome", width=320)
    tree2.column("volume", width=100, anchor="center")

    cache_relatorio = {
        "periodo": ("", ""),
        "linhas": [],  # list of dict
    }

    def _parse_data(txt: str) -> date | None:
        txt = (txt or "").strip()
        if not txt:
            return None
        try:
            return date.fromisoformat(txt)
        except Exception:
            return None

    def gerar():
        d1 = _parse_data(de_var.get())
        d2 = _parse_data(ate_var.get())

        if d1 is None or d2 is None:
            status_var.set("Preencha datas válidas (YYYY-MM-DD).")
            return
        if d2 < d1:
            status_var.set("Período inválido: 'Até' menor que 'De'.")
            return

        movs = listar_movimentos(limite=None)

        # filtra por período (inclui dia inteiro)
        ini_dt = datetime(d1.year, d1.month, d1.day, 0, 0, 0)
        fim_dt = datetime(d2.year, d2.month, d2.day, 23, 59, 59)

        por_item = {}  # nome -> {entradas, saidas, saldo, volume}
        for m in movs:
            dt = _parse_iso_ts(m.get("ts", ""))
            if not dt:
                continue
            if dt < ini_dt or dt > fim_dt:
                continue

            nome = str(m.get("nome", ""))
            delta = _safe_float(m.get("delta", 0), 0.0)

            if nome not in por_item:
                por_item[nome] = {"nome": nome, "entradas": 0.0, "saidas": 0.0, "saldo": 0.0, "volume": 0.0}

            if delta > 0:
                por_item[nome]["entradas"] += delta
            else:
                por_item[nome]["saidas"] += abs(delta)

            por_item[nome]["saldo"] += delta
            por_item[nome]["volume"] += abs(delta)

        linhas = list(por_item.values())
        linhas.sort(key=lambda x: x["nome"].lower())

        tree.delete(*tree.get_children())
        for r in linhas:
            tree.insert("", "end", values=(r["nome"], round(r["entradas"], 3), round(r["saidas"], 3), round(r["saldo"], 3)))

        # top movimentados
        top = sorted(linhas, key=lambda x: x["volume"], reverse=True)
        tree2.delete(*tree2.get_children())
        for r in top[:20]:
            tree2.insert("", "end", values=(r["nome"], round(r["volume"], 3)))

        cache_relatorio["periodo"] = (d1.isoformat(), d2.isoformat())
        cache_relatorio["linhas"] = linhas
        status_var.set(f"{len(linhas)} item(ns) no período")

    def exportar_relatorio_csv():
        linhas = cache_relatorio.get("linhas", [])
        if not linhas:
            messagebox.showwarning("Atenção", "Gere o relatório antes de exportar.")
            return

        de_, ate_ = cache_relatorio.get("periodo", ("", ""))
        nome_sugerido = f"relatorio_{de_}_a_{ate_}.csv".replace("-", "")

        caminho = filedialog.asksaveasfilename(
            title="Salvar relatório em CSV",
            defaultextension=".csv",
            initialfile=nome_sugerido,
            filetypes=[("CSV", "*.csv")],
        )
        if not caminho:
            return

        try:
            with open(caminho, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["item", "entradas", "saidas", "saldo", "volume"])
                for r in linhas:
                    w.writerow([r["nome"], r["entradas"], r["saidas"], r["saldo"], r["volume"]])
            messagebox.showinfo("OK", "Relatório CSV exportado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar CSV:\n\n{e}")

    def exportar_relatorio_excel():
        linhas = cache_relatorio.get("linhas", [])
        if not linhas:
            messagebox.showwarning("Atenção", "Gere o relatório antes de exportar.")
            return

        de_, ate_ = cache_relatorio.get("periodo", ("", ""))
        nome_sugerido = f"relatorio_{de_}_a_{ate_}.xlsx".replace("-", "")

        caminho = filedialog.asksaveasfilename(
            title="Salvar relatório em Excel",
            defaultextension=".xlsx",
            initialfile=nome_sugerido,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not caminho:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"

            header = ["Item", "Entradas", "Saídas", "Saldo", "Volume"]
            ws.append(header)

            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="DDDDDD")
            header_align = Alignment(horizontal="center")

            for col in range(1, len(header) + 1):
                c = ws.cell(row=1, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align

            for r in linhas:
                ws.append([r["nome"], r["entradas"], r["saidas"], r["saldo"], r["volume"]])

            # largura
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

            wb.save(caminho)
            messagebox.showinfo("OK", "Relatório Excel exportado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar Excel:\n\n{e}")

    botoes = ttk.Frame(frame)
    botoes.pack(fill="x", pady=(10, 0))

    ttk.Button(botoes, text="Gerar", command=gerar, takefocus=False).pack(side="right")
    ttk.Button(botoes, text="Exportar Excel", command=exportar_relatorio_excel, takefocus=False).pack(side="right", padx=(0, 8))
    ttk.Button(botoes, text="Exportar CSV", command=exportar_relatorio_csv, takefocus=False).pack(side="right", padx=(0, 8))

    # padrão: últimos 30 dias
    hoje = date.today()
    de_var.set((hoje - timedelta(days=30)).isoformat())
    ate_var.set(hoje.isoformat())
    gerar()


# ============================================================
# 3) AJUSTE DE ESTOQUE (motivo obrigatório na UI)
# ============================================================
def abrir_ajuste_estoque(root: tk.Tk) -> None:
    produtos = carregar_produtos()
    produtos = sorted(produtos, key=lambda p: str(p.get("nome", "")).lower())
    if not produtos:
        messagebox.showinfo("Sem itens", "Cadastre um item antes.")
        return

    # mapa id -> produto
    id_para_produto = {}
    for p in produtos:
        try:
            id_para_produto[int(p.get("id", 0))] = p
        except Exception:
            pass

    win = tk.Toplevel(root)
    win.title("Ajuste de estoque")
    win.geometry("620x520")
    win.minsize(620, 720)
    _configurar_fechamento_toplevel(win, root)
    _singleton_register(root, "relatorios_periodo", win)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Ajuste de estoque", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))
    ttk.Label(frame, text="Use para corrigir o estoque com motivo.").pack(anchor="w", pady=(0, 10))

    # seleção item (reaproveita busca premium)
    ttk.Label(frame, text="Buscar item").pack(anchor="w")
    busca_var = tk.StringVar(value="")
    ent_busca = ttk.Entry(frame, textvariable=busca_var)
    ent_busca.pack(fill="x", pady=(0, 6))

    itens_exibicao = []
    texto_para_id = {}
    item_para_nome_norm = {}
    item_para_tokens = {}
    usados = {}

    for p in produtos:
        nome = str(p.get("nome", "")).strip()
        unidade = str(p.get("unidade", "")).strip()
        pid = int(p.get("id", 0))

        base = nome
        if base in usados:
            base = f"{nome} ({unidade})" if unidade else nome
        n = usados.get(base, 0) + 1
        usados[base] = n
        texto_exibido = base if n == 1 else f"{base} ({n})"

        itens_exibicao.append(texto_exibido)
        texto_para_id[texto_exibido] = pid

        nn = normalizar_busca(nome)
        item_para_nome_norm[texto_exibido] = nn
        item_para_tokens[texto_exibido] = nn.split() if nn else []

    itens_originais = itens_exibicao.copy()

    list_frame = ttk.Frame(frame)
    list_frame.pack(fill="both", expand=True, pady=(0, 10))

    yscroll = ttk.Scrollbar(list_frame, orient="vertical")
    yscroll.pack(side="right", fill="y")

    lb = tk.Listbox(
        list_frame,
        height=10,
        activestyle="none",
        exportselection=False,
        yscrollcommand=yscroll.set,
        bg="#ffffff",
        fg="#1f2933",
        selectbackground="#2563eb",
        selectforeground="#ffffff",
        highlightthickness=1,
        highlightbackground="#d1d5db",
        relief="flat",
        borderwidth=0,
    )
    lb.pack(side="left", fill="both", expand=True)
    yscroll.config(command=lb.yview)

    for it in itens_originais:
        lb.insert("end", it)

    preview_var = tk.StringVar(value="Selecione um item para ver o estoque atual.")
    ttk.Label(frame, textvariable=preview_var).pack(anchor="w", pady=(0, 10))

    def obter_item() -> str | None:
        sel = lb.curselection()
        if not sel:
            return None
        try:
            return lb.get(int(sel[0]))
        except Exception:
            return None

    def atualizar_preview():
        item = obter_item()
        if not item:
            preview_var.set("Selecione um item para ver o estoque atual.")
            return
        pid = texto_para_id.get(item)
        p = id_para_produto.get(int(pid), {})
        un = str(p.get("unidade", "")).strip()
        atual = _safe_float(p.get("estoque_atual", 0), 0.0)
        minimo = _safe_float(p.get("estoque_minimo", 0), 0.0)
        sufixo = f" {un}" if un else ""
        preview_var.set(f"Estoque atual: {atual}{sufixo}  |  Mínimo: {minimo}{sufixo}")

    def filtrar():
        q = normalizar_busca(busca_var.get())
        if not q:
            itens = itens_originais
        else:
            qtoks = q.split()
            a, b, c, d = [], [], [], []
            for item_txt in itens_originais:
                nome_norm = item_para_nome_norm.get(item_txt, "")
                ntoks = item_para_tokens.get(item_txt, [])
                if nome_norm.startswith(q):
                    a.append(item_txt)
                elif match_prefix_por_palavras(qtoks, ntoks):
                    b.append(item_txt)
                elif q in nome_norm:
                    c.append(item_txt)
                elif match_tokens_em_ordem(qtoks, ntoks):
                    d.append(item_txt)
            itens = a + b + c + d

        lb.delete(0, "end")
        for it in itens:
            lb.insert("end", it)

        if len(itens) == 1:
            lb.selection_set(0)
            lb.activate(0)
            lb.see(0)
            atualizar_preview()

    _after = None
    def on_key(event=None):
        nonlocal _after
        if _after is not None:
            try:
                win.after_cancel(_after)
            except Exception:
                pass
        _after = win.after(120, filtrar)

    ent_busca.bind("<KeyRelease>", on_key)
    lb.bind("<<ListboxSelect>>", lambda e: atualizar_preview())

    # motivo
    ttk.Label(frame, text="Motivo do ajuste (obrigatório)").pack(anchor="w")
    motivo_var = tk.StringVar(value="")
    motivos = ["Contagem", "Quebra", "Vencimento", "Doação externa", "Erro de lançamento", "Outro"]
    cb = ttk.Combobox(frame, values=motivos, textvariable=motivo_var, state="readonly")
    cb.pack(fill="x", pady=(0, 10))

    # descrição do "Outro" (aparece só quando motivo == "Outro")
    outro_desc_var = tk.StringVar(value="")

    frm_outro = ttk.Frame(frame)
    ttk.Label(frm_outro, text="Descreva o motivo (obrigatório para 'Outro')").pack(anchor="w")
    ent_outro = ttk.Entry(frm_outro, textvariable=outro_desc_var)
    ent_outro.pack(fill="x")

    def _toggle_outro(*_):
        if motivo_var.get().strip() == "Outro":
            frm_outro.pack(fill="x", pady=(0, 10))
            ent_outro.focus_set()
        else:
            frm_outro.pack_forget()
            outro_desc_var.set("")

    cb.bind("<<ComboboxSelected>>", _toggle_outro)
    # garante estado inicial correto
    _toggle_outro()

    # delta
    ttk.Label(frame, text="Quantidade do ajuste (use negativo para reduzir, ex: -2)").pack(anchor="w")
    qtd_var = tk.StringVar(value="")
    ent_qtd = ttk.Entry(frame, textvariable=qtd_var)
    ent_qtd.pack(fill="x", pady=(0, 10))

    status_var = tk.StringVar(value="")
    lbl_status = ttk.Label(frame, textvariable=status_var)
    lbl_status.pack(anchor="w", pady=(0, 8))

    def aplicar():
        item = obter_item()
        if not item:
            status_var.set("Selecione um item.")
            return
        pid = texto_para_id.get(item)
        if not pid:
            status_var.set("Não foi possível identificar o item.")
            return

        mot = motivo_var.get().strip()
        if not mot:
            status_var.set("Selecione um motivo.")
            cb.focus_set()
            return

    # Se for "Outro", exigir descrição e salvar como "Outro: <descrição>"
        
        if mot == "Outro":
            desc = outro_desc_var.get().strip()
            if not desc:
                status_var.set("Descreva o motivo para 'Outro'.")
                frm_outro.pack(fill="x", pady=(0, 10))
                ent_outro.focus_set()
                return
            mot = f"Outro: {desc}"

        txt = qtd_var.get().strip().replace(",", ".")
        try:
            delta = float(txt)
            if delta == 0:
                raise ValueError()
        except Exception:
            status_var.set("Quantidade inválida. Use um número diferente de 0.")
            ent_qtd.focus_set()
            return

        try:
            atualizar_estoque(int(pid), float(delta), motivo=mot)
        except ValueError as e:
            messagebox.showerror("Erro", str(e))
            ent_qtd.focus_set()
            return

        # recarrega produtos pra preview bater
        novos = carregar_produtos()
        for pp in novos:
            try:
                id_para_produto[int(pp.get("id", 0))] = pp
            except Exception:
                pass

        atualizar_preview()
        status_var.set(f"Ajuste aplicado ({mot}).")
        qtd_var.set("")
        ent_qtd.focus_set()

    ttk.Button(frame, text="Aplicar ajuste", command=aplicar, takefocus=False).pack(anchor="e")


# ============================================================
# 4) INVENTÁRIO / CONTAGEM GUIADA
# ============================================================
def abrir_inventario_contagem(root: tk.Tk) -> None:
    produtos = carregar_produtos()
    produtos = sorted(produtos, key=lambda p: str(p.get("nome", "")).lower())
    if not produtos:
        messagebox.showinfo("Sem itens", "Cadastre um item antes.")
        return

    # janela única
    if _singleton_get(root, "inventario") is not None:
        return

    win = tk.Toplevel(root)
    win.title("Inventário / Contagem")
    win.state("zoomed")  # tela cheia no Windows
    win.minsize(1100, 720)
    _configurar_fechamento_toplevel(win, root)
    _singleton_register(root, "inventario", win)

    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Inventário / Contagem guiada", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))
    ttk.Label(frame, text="Digite o 'Contado' e aplique para ajustar o estoque ao valor real.").pack(anchor="w", pady=(0, 10))

    topo = ttk.Frame(frame)
    topo.pack(fill="x", pady=(0, 10))

    ttk.Label(topo, text="Filtrar item").pack(side="left")
    filtro_var = tk.StringVar(value="")
    ent_filtro = ttk.Entry(topo, textvariable=filtro_var, width=30)
    ent_filtro.pack(side="left", padx=(8, 16))

    status_var = tk.StringVar(value="")
    ttk.Label(topo, textvariable=status_var).pack(side="right")

    cols = ("id", "nome", "un", "atual", "contado", "diff")
    tree = ttk.Treeview(frame, columns=cols, show="headings", height=18)
    tree.pack(fill="both", expand=True)

    # mostra tudo internamente, mas oculta visualmente o ID (mantém funcionamento normal)
    tree["displaycolumns"] = ("nome", "un", "atual", "contado", "diff")

    # Ordenação por clique no cabeçalho (profissional)
    _sort_inv = {"col": None, "desc": False}

    def _coerce_inv(col: str, v: str):
        if col in ("id", "atual", "contado", "diff"):
            try:
                return float(str(v).replace(",", "."))
            except Exception:
                return 0.0
        return str(v).lower()

    def ordenar_inv(col: str):
        if _sort_inv["col"] == col:
            _sort_inv["desc"] = not _sort_inv["desc"]
        else:
            _sort_inv["col"] = col
            _sort_inv["desc"] = False

        itens = list(tree.get_children(""))
        itens.sort(key=lambda iid: _coerce_inv(col, tree.set(iid, col)), reverse=_sort_inv["desc"])
        for i, iid in enumerate(itens):
            tree.move(iid, "", i)

    tree.heading("id", text="ID", command=lambda: ordenar_inv("id"))
    tree.heading("nome", text="Item", command=lambda: ordenar_inv("nome"))
    tree.heading("un", text="Unid.", command=lambda: ordenar_inv("un"))
    tree.heading("atual", text="Atual", command=lambda: ordenar_inv("atual"))
    tree.heading("contado", text="Contado", command=lambda: ordenar_inv("contado"))
    tree.heading("diff", text="Diferença", command=lambda: ordenar_inv("diff"))

    tree.column("id", width=60, anchor="center")
    tree.column("nome", width=380)
    tree.column("un", width=70, anchor="center")
    tree.column("atual", width=90, anchor="center")
    tree.column("contado", width=90, anchor="center")
    tree.column("diff", width=90, anchor="center")

    # Editor visível (mais óbvio que dá pra digitar o "Contado")
    editor_frame = ttk.Frame(frame)
    editor_frame.pack(fill="x", pady=(8, 0))

    selecionado_var = tk.StringVar(value="Selecione um item na tabela para informar o contado.")
    ttk.Label(editor_frame, textvariable=selecionado_var).pack(side="left")

    ttk.Label(editor_frame, text="Contado").pack(side="left", padx=(16, 6))
    contado_var = tk.StringVar(value="")
    ent_contado = ttk.Entry(editor_frame, textvariable=contado_var, width=12)
    ent_contado.pack(side="left")

    def _aplicar_contado_selecionado():
        sel = tree.selection()
        if not sel:
            status_var.set("Selecione um item na tabela.")
            return
        iid = sel[0]
        txt = contado_var.get().strip().replace(",", ".")
        try:
            v = float(txt)
            if v < 0:
                raise ValueError()
        except Exception:
            status_var.set("Contado inválido. Use um número >= 0 (ex: 10 ou 10,5).")
            ent_contado.focus_set()
            return

        r = linhas.get(iid)
        if not r:
            status_var.set("Item selecionado inválido.")
            return

        r["contado"] = float(v)
        atual = float(r.get("atual", 0.0))
        diff = float(v) - atual
        tree.set(iid, "contado", float(v))
        tree.set(iid, "diff", diff)

        status_var.set("Contado atualizado na tabela. Clique em 'Aplicar ajustes' para gravar no estoque.")
        try:
            ent_contado.selection_range(0, "end")
            ent_contado.focus_set()
        except Exception:
            pass

    ttk.Button(editor_frame, text="Aplicar no selecionado", command=_aplicar_contado_selecionado, takefocus=False).pack(side="left", padx=(8, 0))

    def _on_select_row(_e=None):
        sel = tree.selection()
        if not sel:
            selecionado_var.set("Selecione um item na tabela para informar o contado.")
            return
        iid = sel[0]
        r = linhas.get(iid, {})
        nome = r.get("nome", "")
        atual = r.get("atual", "")
        selecionado_var.set(f"Selecionado: {nome} | Atual: {atual}")
        # prepara campo
        cv = r.get("contado")
        contado_var.set("" if cv is None else str(cv))

    tree.bind("<<TreeviewSelect>>", _on_select_row)


    # cache: iid -> dict com valores
    linhas: dict[str, dict] = {}

    def render():
        tree.delete(*tree.get_children())
        linhas.clear()

        termo = normalizar_busca(filtro_var.get().strip())
        count = 0

        for p in produtos:
            nome = str(p.get("nome", ""))
            if termo and termo not in normalizar_busca(nome):
                continue

            pid = int(p.get("id", 0))
            un = str(p.get("unidade", "")).strip()
            atual = _safe_float(p.get("estoque_atual", 0), 0.0)
            contado = ""  # vazio inicialmente
            diff = ""

            iid = tree.insert("", "end", values=(pid, nome, un, atual, contado, diff))
            linhas[iid] = {"id": pid, "nome": nome, "un": un, "atual": atual, "contado": None}
            count += 1

        status_var.set(f"{count} item(ns)")

    def _recalc(iid: str):
        r = linhas.get(iid)
        if not r:
            return
        atual = float(r.get("atual", 0.0))
        contado = r.get("contado")
        if contado is None:
            tree.set(iid, "diff", "")
            return
        try:
            diff = float(contado) - atual
        except Exception:
            tree.set(iid, "diff", "")
            return
        tree.set(iid, "diff", round(diff, 3))

    # editor simples de célula (contado)
    editor = {"w": None, "iid": None}

    def _close_editor():
        w = editor.get("w")
        if w is not None:
            try:
                w.destroy()
            except Exception:
                pass
        editor["w"] = None
        editor["iid"] = None

    def _start_edit(event=None):
        # só edita coluna "contado"
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)  # "#1"...
        if col != "#5":  # 5a coluna = contado
            return

        iid = tree.identify_row(event.y)
        if not iid:
            return

        bbox = tree.bbox(iid, "contado")
        if not bbox:
            return

        x, y, w, h = bbox
        value = tree.set(iid, "contado")

        _close_editor()

        e = ttk.Entry(tree)
        e.place(x=x, y=y, width=w, height=h)
        e.insert(0, value)
        e.focus_set()

        editor["w"] = e
        editor["iid"] = iid

        def _commit(_evt=None):
            txt = e.get().strip().replace(",", ".")
            if txt == "":
                tree.set(iid, "contado", "")
                linhas[iid]["contado"] = None
                _recalc(iid)
                _close_editor()
                return "break"

            try:
                val = float(txt)
                if val < 0:
                    raise ValueError()
            except Exception:
                # não trava app: só não aplica
                tree.set(iid, "contado", "")
                linhas[iid]["contado"] = None
                _recalc(iid)
                _close_editor()
                return "break"

            tree.set(iid, "contado", str(val))
            linhas[iid]["contado"] = float(val)
            _recalc(iid)
            _close_editor()
            return "break"

        e.bind("<Return>", _commit)
        e.bind("<Escape>", lambda _e: (_close_editor(), "break"))
        e.bind("<FocusOut>", _commit)

    tree.bind("<Double-Button-1>", _start_edit)

    def aplicar_ajustes():
        # motivo fixo (UI): "Contagem"
        mot = "Contagem"
        ajustes = []
        for iid, r in linhas.items():
            contado = r.get("contado")
            if contado is None:
                continue
            atual = float(r.get("atual", 0.0))
            delta = float(contado) - atual
            if abs(delta) < 1e-9:
                continue
            ajustes.append((r["id"], delta))

        if not ajustes:
            messagebox.showinfo("OK", "Nada para ajustar.")
            return

        # aplica em lote
        erros = 0
        for pid, delta in ajustes:
            try:
                atualizar_estoque(int(pid), float(delta), motivo=mot)
            except Exception:
                erros += 1

        # recarrega produtos e re-renderiza
        nonlocal_prod = carregar_produtos()
        nonlocal_prod = sorted(nonlocal_prod, key=lambda p: str(p.get("nome", "")).lower())
        produtos[:] = nonlocal_prod  # mantém referência
        render()

        if erros:
            messagebox.showwarning("Concluído", f"Ajustes aplicados com {erros} erro(s).")
        else:
            messagebox.showinfo("Concluído", f"Ajustes aplicados: {len(ajustes)}")

    def exportar_lista_contagem():
        # exporta lista com Atual e campo Contado vazio (para imprimir/usar fora)
        nome_sugerido = f"lista_contagem_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        caminho = filedialog.asksaveasfilename(
            title="Salvar lista de contagem (CSV)",
            defaultextension=".csv",
            initialfile=nome_sugerido,
            filetypes=[("CSV", "*.csv")],
        )
        if not caminho:
            return

        try:
            termo = normalizar_busca(filtro_var.get().strip())
            with open(caminho, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["id", "item", "unidade", "estoque_atual", "contado"])
                for p in produtos:
                    nome = str(p.get("nome", ""))
                    if termo and termo not in normalizar_busca(nome):
                        continue
                    w.writerow([p.get("id", ""), nome, p.get("unidade", ""), p.get("estoque_atual", 0), ""])
            messagebox.showinfo("OK", "Lista de contagem exportada.")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar:\n\n{e}")

    botoes = ttk.Frame(frame)
    botoes.pack(fill="x", pady=(10, 0))

    ttk.Button(botoes, text="Exportar lista (CSV)", command=exportar_lista_contagem, takefocus=False).pack(side="left")
    ttk.Button(botoes, text="Aplicar ajustes", command=aplicar_ajustes, takefocus=False).pack(side="right")

    ent_filtro.bind("<KeyRelease>", lambda e: render())
    render()


def ajustar_janela_ao_conteudo_e_centralizar(root: tk.Tk, margem: int = 24) -> None:
    root.update_idletasks()
    w = root.winfo_reqwidth() + margem
    h = root.winfo_reqheight() + margem

    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x = (sw - w) // 2
    y = (sh - h) // 2

    root.geometry(f"{w}x{h}+{x}+{y}")

def centralizar_janela(root: tk.Tk) -> None:
    root.update_idletasks()

    w = root.winfo_width()
    h = root.winfo_height()

    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()

    x = (sw - w) // 2
    y = (sh - h) // 2

    root.geometry(f"{w}x{h}+{x}+{y}")


# =========================
# UX: "Hubs" para reduzir poluição do menu inicial
# =========================
def _abrir_hub(root: tk.Tk, titulo: str, subtitulo: str | None, acoes: list[tuple[str, callable]]) -> None:
    """
    Abre uma janela simples com botões para agrupar ações.

    Melhorias de UX (apenas para hubs):
    - Evita múltiplos hubs abertos: ao abrir um hub, fecha outros hubs.
    - Navegação linear: ao clicar numa ação do hub, fecha o hub e abre a tela destino.
    - Usa grab_set() para impedir interação com o root por trás (sem fechar o app).
    - Evita "flash" (canto superior esquerdo): cria oculto (withdraw) e exibe após configurar.
    """
    win = tk.Toplevel(root)
    win.withdraw()  # evita o flash antes de posicionar
    win._is_hub = True  # marcador: esta janela é um hub

    # Fecha outros hubs abertos (não fecha telas de trabalho)
    for w in root.winfo_children():
        if w is win:
            continue
        if isinstance(w, tk.Toplevel) and getattr(w, "_is_hub", False):
            try:
                w.destroy()
            except Exception:
                pass

    win.title(titulo)
    win.geometry("520x420")
    win.minsize(520, 200)

    def _on_close() -> None:
        # Libera o grab (se houver) para não "travar" o app
        try:
            win.grab_release()
        except Exception:
            pass
        try:
            win.destroy()
        finally:
            root.after(0, lambda: _restaurar_foco_menu(root))

    win.protocol("WM_DELETE_WINDOW", _on_close)

    frame = ttk.Frame(win, padding=16)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text=titulo, font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 6))
    if subtitulo:
        ttk.Label(frame, text=subtitulo).pack(anchor="w", pady=(0, 12))

    def _ir_para(fn: callable) -> None:
        # Fecha o hub antes de abrir a próxima tela
        try:
            win.grab_release()
        except Exception:
            pass
        try:
            win.destroy()
        except Exception:
            pass
        root.after(0, fn)

    for label, fn in acoes:
        ttk.Button(
            frame,
            text=label,
            style="Menu.TButton",
            takefocus=False,
            command=lambda f=fn: _ir_para(f),
        ).pack(fill="x", pady=6)

    ttk.Button(
        frame,
        text="Fechar",
        takefocus=False,
        command=_on_close,
    ).pack(anchor="e", pady=(14, 0))

    # Só exibe depois de posicionar e montar widgets (evita flash)
    win.update_idletasks()
    centralizar_janela(win)
    win.deiconify()

    # Garante que o hub fique na frente e capture interação
    try:
        win.lift()
        win.focus_force()
        win.grab_set()
    except Exception:
        pass


def abrir_hub_movimentar(root: tk.Tk) -> None:
    _abrir_hub(
        root,
        titulo="Movimentar estoque",
        subtitulo="Registre entradas e saídas de forma contínua.",
        acoes=[
            ("Entrada", lambda: abrir_movimento(root, "entrada")),
            ("Saída", lambda: abrir_movimento(root, "saida")),
        ],
    )


def abrir_hub_consultar(root: tk.Tk) -> None:
    # Informação “itens em falta” aqui (sai do menu inicial)
    try:
        produtos = carregar_produtos()
        abaixo = [
            p for p in produtos
            if _safe_float(p.get("estoque_atual", 0), 0.0) < _safe_float(p.get("estoque_minimo", 0), 0.0)
        ]
        info = f"Itens em falta (abaixo do mínimo): {len(abaixo)}"
    except Exception:
        info = "Consultar itens e visão geral do estoque."

    _abrir_hub(
        root,
        titulo="Consultar",
        subtitulo=info,
        acoes=[
            ("Visão Geral", lambda: abrir_dashboard(root)),
            ("Estoque", lambda: abrir_tela_produtos(root)),
        ],
    )


def abrir_hub_movimentacoes(root: tk.Tk) -> None:
    _abrir_hub(
        root,
        titulo="Movimentações",
        subtitulo="Histórico, relatórios por período e inventário (contagem).",
        acoes=[
            ("Histórico", lambda: abrir_historico(root)),
            ("Relatórios (período)", lambda: abrir_relatorios_periodo(root)),
            ("Inventário / Contagem", lambda: abrir_inventario_contagem(root)),
            ("Ajuste de estoque", lambda: abrir_ajuste_estoque(root)),
        ],
    )

def main():
    root = tk.Tk()
    root.withdraw()  # evita o "flash" e aparecer em outro lugar por milissegundos

    # ===== TEMA MODERNO (padrão LIGHT) =====
    # (Opcional) sv-ttk deixa a interface com cara mais moderna.
    # Se não estiver instalado, o app roda normalmente com ttk padrão.
    try:
        import sv_ttk  # type: ignore
        TEMA_OK = True
    except Exception:
        TEMA_OK = False

    if TEMA_OK:
        sv_ttk.set_theme("light")

    # ===== ESTILO GLOBAL (só visual) =====
    style = ttk.Style()

    # Fonte moderna padrão do Windows (com espaço no nome)
    root.option_add("*Font", "{Segoe UI} 10")

    # Botões com mais espaçamento (mais “app moderno”)
    style.configure("TButton", padding=(12, 10))
    style.configure("Menu.TButton", padding=(12, 10))

    # Tabelas (Treeview) mais elegantes
    style.configure("Treeview", rowheight=28)
    style.configure("Treeview.Heading", font=("{Segoe UI}", 10, "bold"))

    # ===== TÍTULO + ÍCONE =====
    root.title("Estoque – Vila Vicentina de Abaeté")
    if ICONE_ICO.exists():
        try:
            root.iconbitmap(str(ICONE_ICO))
        except Exception:
            pass


    # ===== MENU DE TEMA (dark/light) =====
    # Só aparece se o sv-ttk estiver instalado
    if TEMA_OK:
        menubar = tk.Menu(root)
        visual = tk.Menu(menubar, tearoff=0)

        def usar_dark():
            sv_ttk.set_theme("dark")

        def usar_light():
            sv_ttk.set_theme("light")

        visual.add_command(label="Tema claro", command=usar_light)
        visual.add_command(label="Tema escuro", command=usar_dark)
        menubar.add_cascade(label="Visual", menu=visual)
        root.config(menu=menubar)

    # ===== LAYOUT PRINCIPAL =====
    frame = ttk.Frame(root, padding=16)
    frame.pack(fill="both", expand=True)

    # guarda o widget que deve receber foco quando voltar das telas
    root._menu_focus_widget = frame
    frame.configure(takefocus=True)
    frame.focus_set()

    ttk.Label(
        frame,
        text="Controle de Estoque – Vila Vicentina",
        font=("{Segoe UI}", 15, "bold"),
    ).pack(pady=(0, 12))


    # ===== MENU INICIAL (ENXUTO) =====
    ttk.Button(
        frame,
        text="Movimentar Estoque",
        style="Menu.TButton",
        takefocus=False,
        command=lambda: abrir_hub_movimentar(root),
    ).pack(fill="x", pady=6)

    ttk.Button(
        frame,
        text="Consultar",
        style="Menu.TButton",
        takefocus=False,
        command=lambda: abrir_hub_consultar(root),
    ).pack(fill="x", pady=6)

    ttk.Button(
        frame,
        text="Movimentações",
        style="Menu.TButton",
        takefocus=False,
        command=lambda: abrir_hub_movimentacoes(root),
    ).pack(fill="x", pady=6)

    ttk.Button(
        frame,
        text="Cadastrar item",
        style="Menu.TButton",
        takefocus=False,
        command=lambda: abrir_cadastro_produto(root),
    ).pack(fill="x", pady=6)

    ajustar_janela_ao_conteudo_e_centralizar(root)

    root.deiconify()  # mostra só quando já está pronto e centralizado
    root.mainloop()